from openpyxl import load_workbook

def read_login_data():
    wb = load_workbook("login_data.xlsx")
    sheet = wb["Sheet1"]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data