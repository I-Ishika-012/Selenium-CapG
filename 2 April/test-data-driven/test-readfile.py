import pytest
# import openpyxl
# wb = openpyxl.load_workbook("sauce.xlsx")
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

def read_login_data():
    wb = load_workbook("sauce.xlsx")
    sheet = wb["Sheet1"]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data


@pytest.mark.parametrize("uname, passwd", read_login_data())
def test_login(uname, passwd):

    driver = webdriver.Chrome()
    driver.get("https://www.saucedemo.com/")

    driver.find_element(By.ID, "user-name").send_keys(uname)
    driver.find_element(By.ID, "password").send_keys(passwd)
    driver.find_element(By.ID, "login-button").click()

    assert "inventory" in driver.current_url

    driver.quit()